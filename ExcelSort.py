from openpyxl import load_workbook
from datetime import datetime
from datetime import timedelta
from calendar import month_abbr
Today = datetime.now()
Yesterday = (Today - timedelta(days = 1))
Year = Today.strftime("%Y")


class ExcelOrganizer():
    '''Class Function used to organize Excel Sheets by Month'''
    def __init__(self,Filename):
        self.filename = Filename

    def organizefile(self):
        '''Rename the sheet names by month name and then combine the other sheets'''
        try: 
            wb = load_workbook(filename=f'{self.filename}')
            for i in wb._sheets:
                if i.title.lower().startswith('jan'):
                    i.title = 'Jan'
                elif i.title.lower().startswith('feb'):
                    i.title = 'Feb'
                elif i.title.lower().startswith('mar'):
                    i.title = 'Mar'
                elif i.title.lower().startswith('apr'):
                    i.title = 'Apr'
                elif i.title.lower().startswith('may'):
                    i.title = 'May'
                elif i.title.lower().startswith('jun'):
                    i.title = 'Jun'
                elif i.title.lower().startswith('jul'):
                    i.title = 'Jul'
                elif i.title.lower().startswith('aug'):
                    i.title = 'Aug'
                elif i.title.lower().startswith('sep'):
                    i.title = 'Sep'
                elif i.title.lower().startswith('oct'):
                    i.title = 'Oct'
                elif i.title.lower().startswith('nov'):
                    i.title = 'Nov'
                elif i.title.lower().startswith('dec'):
                    i.title = 'Dec'
                else:
                    i.title = i.title 
            Months = list(month_abbr)
            SheetOrder = {val: idx for idx, val in enumerate(Months)}
            CurrentOrder = {val: idx for idx, val in enumerate(wb.sheetnames)}
            NewOrder = CurrentOrder | SheetOrder
            wb._sheets.sort(key=lambda x: NewOrder[x.title])
            wb.save(filename=f'{self.filename}')
            return True
        except:
            return False
